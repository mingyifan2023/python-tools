# -*- coding: utf_8 -*-


import json
import openpyxl
import pandas as pd
import sys
import csv
from os.path import dirname, abspath, join
# 获取当前文件所在的绝对路径的上级目录
project_dir = dirname(dirname(abspath(__file__)))

# 将templates目录添加到sys.path中
common_path = join(project_dir, 'common')
if common_path not in sys.path:
    sys.path.append(common_path)




class FileExec():
    # 将 DataFrame 中的数据转换为 utf-8 编码
    def encode_utf8(self,x):
        try:
            return str(x).encode('utf-8').decode('utf-8')
        except AttributeError:
            return x

    def execl_to_json_with_pandas(self,xlsxfile,jsonfile):
        # 读取 Excel 文件
        # 打开CSV文件并指定编码格式为GBK
        # 读取Excel文件为DataFrame
        df = pd.read_excel(xlsxfile)

        data_list = []

        # 遍历DataFrame的每一行数据，将每行数据转换为字典并添加到列表中
        for index, row in df.iterrows():
            data_dict = {
                "name": row['name'],
                "code": row['code'],
                "type1": row['type1'],
                "type2": row['type2'],
                "weight": row['weight']
            }
            data_list.append(data_dict)

        # 将包含字典的列表写入JSON文件
        with open(jsonfile, 'w', encoding='utf-8') as file:
            json.dump(data_list, file, ensure_ascii=False, indent=4)



    def execl_to_json_with_openxyl(self,xlsxfile,jsonfile):
        # 打开Excel文件
        wb = openpyxl.load_workbook(xlsxfile)

        # 选择要读取的sheet
        sheet = wb.active

        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append([str(cell) for cell in row])

        df = pd.DataFrame(data)

        # 指定编码格式为UTF-8
        df = df.applymap(lambda x: x.encode('latin1').decode('utf-8'))

        print(df)

        # # 读取 Excel 文件
        # wb = openpyxl.load_workbook(xlsxfile)
        # sheet = wb.active
        #
        # # 将 Excel 数据转换为字典
        # data = []
        # for row in sheet.iter_rows(values_only=True):
        #     data.append(row)
        #
        # # 获取表头作为键
        # keys = data[0]
        # json_data = [dict(zip(keys, row)) for row in data[1:]]
        #
        # # 将数据输出为 JSON 文件
        # with open(jsonfile, 'w') as f:
        #     json.dump(json_data, f, ensure_ascii=False)

    def readjsonfile(self,filename):
        with open(filename, 'r', encoding='utf-8') as fw:
            ret = json.load(fw)
            return ret

f = FileExec()
xlsxfile = join(common_path,"nikkei225.xlsx")
f.execl_to_json_with_pandas(xlsxfile,"nikkei225.json")
